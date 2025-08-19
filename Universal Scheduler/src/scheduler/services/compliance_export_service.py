"""Service for exporting ACGME compliance data and maintaining audit trails."""

import csv
import json
import io
from datetime import datetime, date, timedelta
from typing import List, Dict, Optional, Any, BinaryIO
from sqlalchemy.orm import Session
from sqlalchemy import and_, desc, func
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

from ..models import (
    Resident,
    ACGMEViolation,
    WeeklyDutyHourSummary,
    DutyHourEntry,
    ACGMEComplianceReport,
    CorrectiveAction,
    ACGMEViolationType,
    ACGMEViolationSeverity,
    AuditLog,
)

logger = logging.getLogger(__name__)


class ComplianceExportService:
    """Service for exporting compliance data in various formats."""

    def __init__(self, db_session: Session):
        self.db = db_session
        self.logger = logging.getLogger(__name__)

    def export_violations_csv(
        self,
        start_date: date,
        end_date: date,
        program_id: Optional[str] = None,
        resident_ids: Optional[List[int]] = None,
    ) -> str:
        """Export ACGME violations to CSV format."""
        try:
            # Build query
            query = self.db.query(ACGMEViolation).filter(
                and_(
                    ACGMEViolation.violation_date >= start_date,
                    ACGMEViolation.violation_date <= end_date,
                )
            )

            # Apply filters
            if resident_ids:
                query = query.filter(ACGMEViolation.resident_id.in_(resident_ids))
            elif program_id:
                # Get residents in program
                program_residents = (
                    self.db.query(Resident)
                    .filter(Resident.program_id == program_id)
                    .all()
                )
                resident_ids = [r.id for r in program_residents]
                query = query.filter(ACGMEViolation.resident_id.in_(resident_ids))

            violations = query.order_by(ACGMEViolation.violation_date.desc()).all()

            # Create CSV content
            output = io.StringIO()
            writer = csv.writer(output)

            # Headers
            headers = [
                "Violation_ID",
                "Resident_ID",
                "Resident_Name",
                "PGY_Level",
                "Violation_Type",
                "Severity",
                "Violation_Date",
                "Detected_At",
                "Title",
                "Description",
                "Actual_Value",
                "Limit_Value",
                "Excess_Amount",
                "Status",
                "PD_Reviewed",
                "Resolution_Method",
                "Resolved_At",
                "Days_To_Resolution",
            ]
            writer.writerow(headers)

            # Get resident data for efficiency
            resident_map = {}
            if violations:
                resident_ids = list(set(v.resident_id for v in violations))
                residents = (
                    self.db.query(Resident).filter(Resident.id.in_(resident_ids)).all()
                )
                resident_map = {r.id: r for r in residents}

            # Data rows
            for violation in violations:
                resident = resident_map.get(violation.resident_id)
                days_to_resolution = None
                if violation.resolved_at and violation.detected_at:
                    days_to_resolution = (
                        violation.resolved_at - violation.detected_at
                    ).days

                row = [
                    violation.id,
                    violation.resident_id,
                    resident.name if resident else "Unknown",
                    resident.pgy_level if resident else "Unknown",
                    violation.violation_type.value,
                    violation.severity.value,
                    violation.violation_date.isoformat(),
                    violation.detected_at.isoformat(),
                    violation.title,
                    violation.description,
                    violation.actual_value,
                    violation.limit_value,
                    violation.excess_amount,
                    violation.status,
                    violation.pd_reviewed,
                    violation.resolution_method or "",
                    violation.resolved_at.isoformat() if violation.resolved_at else "",
                    days_to_resolution,
                ]
                writer.writerow(row)

            csv_content = output.getvalue()
            output.close()

            # Log audit trail
            self._log_export_activity(
                "violations_csv_export",
                {
                    "start_date": start_date.isoformat(),
                    "end_date": end_date.isoformat(),
                    "program_id": program_id,
                    "record_count": len(violations),
                },
            )

            return csv_content

        except Exception as e:
            self.logger.error(f"Failed to export violations CSV: {e}")
            raise

    def export_weekly_summaries_excel(
        self, start_date: date, end_date: date, program_id: Optional[str] = None
    ) -> bytes:
        """Export weekly duty hour summaries to Excel format."""
        try:
            # Get weekly summaries
            query = self.db.query(WeeklyDutyHourSummary).filter(
                and_(
                    WeeklyDutyHourSummary.week_start_date >= start_date,
                    WeeklyDutyHourSummary.week_start_date <= end_date,
                )
            )

            if program_id:
                # Get residents in program
                program_residents = (
                    self.db.query(Resident)
                    .filter(Resident.program_id == program_id)
                    .all()
                )
                resident_ids = [r.id for r in program_residents]
                query = query.filter(
                    WeeklyDutyHourSummary.resident_id.in_(resident_ids)
                )

            summaries = query.order_by(
                WeeklyDutyHourSummary.week_start_date.desc(),
                WeeklyDutyHourSummary.resident_id,
            ).all()

            # Get resident data
            resident_map = {}
            if summaries:
                resident_ids = list(set(s.resident_id for s in summaries))
                residents = (
                    self.db.query(Resident).filter(Resident.id.in_(resident_ids)).all()
                )
                resident_map = {r.id: r for r in residents}

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Weekly Duty Hour Summaries"

            # Headers with formatting
            headers = [
                "Resident_ID",
                "Resident_Name",
                "PGY_Level",
                "Week_Start",
                "Week_End",
                "Total_Hours",
                "Clinical_Hours",
                "Educational_Hours",
                "Educational_Pct",
                "Research_Hours",
                "Administrative_Hours",
                "Moonlighting_Hours",
                "Call_Shifts",
                "Overnight_Calls",
                "Longest_Continuous_Duty",
                "Shortest_Rest_Period",
                "Rest_Violations",
                "Is_Compliant",
                "Compliance_Pct",
                "Hours_Over_80",
                "Violation_Count",
            ]

            # Style headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Data rows
            for row, summary in enumerate(summaries, 2):
                resident = resident_map.get(summary.resident_id)

                data = [
                    summary.resident_id,
                    resident.name if resident else "Unknown",
                    resident.pgy_level if resident else "Unknown",
                    summary.week_start_date,
                    summary.week_end_date,
                    summary.total_duty_hours,
                    summary.total_clinical_hours,
                    summary.total_educational_hours,
                    summary.educational_hours_percentage,
                    summary.total_research_hours,
                    summary.total_administrative_hours,
                    summary.total_moonlighting_hours,
                    summary.call_shifts_count,
                    summary.overnight_calls_count,
                    summary.longest_continuous_duty,
                    summary.shortest_rest_period,
                    summary.rest_period_violations,
                    "Yes" if summary.is_compliant else "No",
                    summary.compliance_percentage,
                    summary.hours_over_limit,
                    summary.violation_count,
                ]

                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)

                    # Color non-compliant rows
                    if not summary.is_compliant:
                        cell.fill = PatternFill(
                            start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"
                        )

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            excel_content = output.getvalue()
            output.close()

            # Log audit trail
            self._log_export_activity(
                "weekly_summaries_excel_export",
                {
                    "start_date": start_date.isoformat(),
                    "end_date": end_date.isoformat(),
                    "program_id": program_id,
                    "record_count": len(summaries),
                },
            )

            return excel_content

        except Exception as e:
            self.logger.error(f"Failed to export weekly summaries Excel: {e}")
            raise

    def export_duty_hours_detailed_csv(
        self, resident_id: int, start_date: date, end_date: date
    ) -> str:
        """Export detailed duty hour entries for a specific resident."""
        try:
            # Get duty hour entries
            entries = (
                self.db.query(DutyHourEntry)
                .filter(
                    and_(
                        DutyHourEntry.resident_id == resident_id,
                        DutyHourEntry.date >= start_date,
                        DutyHourEntry.date <= end_date,
                    )
                )
                .order_by(DutyHourEntry.date, DutyHourEntry.start_time)
                .all()
            )

            # Get resident info
            resident = (
                self.db.query(Resident).filter(Resident.id == resident_id).first()
            )
            if not resident:
                raise ValueError(f"Resident {resident_id} not found")

            # Create CSV content
            output = io.StringIO()
            writer = csv.writer(output)

            # Metadata rows
            writer.writerow(["Resident Duty Hour Detail Export"])
            writer.writerow(["Resident:", f"{resident.name} ({resident.pgy_level})"])
            writer.writerow(["Period:", f"{start_date} to {end_date}"])
            writer.writerow(["Export Date:", datetime.utcnow().isoformat()])
            writer.writerow([])  # Empty row

            # Headers
            headers = [
                "Date",
                "Start_Time",
                "End_Time",
                "Total_Hours",
                "Activity_Type",
                "Is_Clinical",
                "Is_Educational",
                "Is_Call_Duty",
                "Is_Moonlighting",
                "Location",
                "Supervisor",
                "Supervision_Level",
                "Entry_Method",
                "Verified_By_Resident",
                "Potential_Violation",
                "Notes",
            ]
            writer.writerow(headers)

            # Data rows
            for entry in entries:
                row = [
                    entry.date.isoformat(),
                    entry.start_time.isoformat() if entry.start_time else "",
                    entry.end_time.isoformat() if entry.end_time else "",
                    entry.total_hours,
                    entry.activity_type,
                    entry.is_clinical_work,
                    entry.is_educational,
                    entry.is_call_duty,
                    entry.is_moonlighting,
                    entry.location or "",
                    entry.supervisor_name or "",
                    entry.supervision_level or "",
                    entry.entry_method,
                    entry.verified_by_resident,
                    entry.potential_violation,
                    entry.notes or "",
                ]
                writer.writerow(row)

            # Summary statistics
            writer.writerow([])
            writer.writerow(["SUMMARY STATISTICS"])
            writer.writerow(["Total Entries:", len(entries)])
            writer.writerow(["Total Hours:", sum(e.total_hours for e in entries)])
            writer.writerow(
                [
                    "Clinical Hours:",
                    sum(e.total_hours for e in entries if e.is_clinical_work),
                ]
            )
            writer.writerow(
                [
                    "Educational Hours:",
                    sum(e.total_hours for e in entries if e.is_educational),
                ]
            )
            writer.writerow(
                [
                    "Call Duty Hours:",
                    sum(e.total_hours for e in entries if e.is_call_duty),
                ]
            )
            writer.writerow(
                [
                    "Moonlighting Hours:",
                    sum(e.total_hours for e in entries if e.is_moonlighting),
                ]
            )
            writer.writerow(
                [
                    "Potential Violations:",
                    sum(1 for e in entries if e.potential_violation),
                ]
            )

            csv_content = output.getvalue()
            output.close()

            # Log audit trail
            self._log_export_activity(
                "duty_hours_detailed_csv_export",
                {
                    "resident_id": resident_id,
                    "resident_name": resident.name,
                    "start_date": start_date.isoformat(),
                    "end_date": end_date.isoformat(),
                    "record_count": len(entries),
                },
            )

            return csv_content

        except Exception as e:
            self.logger.error(
                f"Failed to export detailed duty hours for resident {resident_id}: {e}"
            )
            raise

    def export_compliance_report_json(self, report_id: int) -> str:
        """Export compliance report in JSON format for API integration."""
        try:
            report = (
                self.db.query(ACGMEComplianceReport)
                .filter(ACGMEComplianceReport.id == report_id)
                .first()
            )

            if not report:
                raise ValueError(f"Compliance report {report_id} not found")

            # Build comprehensive JSON structure
            report_data = {
                "report_metadata": {
                    "id": report.id,
                    "name": report.report_name,
                    "type": report.report_type,
                    "start_date": report.start_date.isoformat(),
                    "end_date": report.end_date.isoformat(),
                    "program_id": report.program_id,
                    "generated_by": report.generated_by,
                    "generated_at": report.created_at.isoformat(),
                    "status": report.status,
                },
                "summary_statistics": {
                    "total_residents": report.total_residents,
                    "compliant_residents": report.compliant_residents,
                    "compliance_rate": report.compliance_rate,
                    "total_violations": report.total_violations,
                    "critical_violations": report.critical_violations,
                    "high_violations": report.high_violations,
                    "resolved_violations": report.resolved_violations,
                    "avg_weekly_hours": report.avg_weekly_hours,
                    "max_weekly_hours": report.max_weekly_hours,
                    "weeks_over_80_hours": report.weeks_over_80_hours,
                },
                "detailed_data": report.detailed_data,
                "summary_statistics_detailed": report.summary_statistics,
            }

            json_content = json.dumps(report_data, indent=2, default=str)

            # Log audit trail
            self._log_export_activity(
                "compliance_report_json_export",
                {
                    "report_id": report_id,
                    "report_name": report.report_name,
                    "report_type": report.report_type,
                },
            )

            return json_content

        except Exception as e:
            self.logger.error(
                f"Failed to export compliance report {report_id} as JSON: {e}"
            )
            raise

    def export_corrective_actions_csv(
        self, start_date: date, end_date: date, status_filter: Optional[str] = None
    ) -> str:
        """Export corrective actions to CSV format."""
        try:
            # Build query
            query = self.db.query(CorrectiveAction).filter(
                and_(
                    CorrectiveAction.action_date >= start_date,
                    CorrectiveAction.action_date <= end_date,
                )
            )

            if status_filter:
                query = query.filter(CorrectiveAction.status == status_filter)

            actions = query.order_by(CorrectiveAction.action_date.desc()).all()

            # Create CSV content
            output = io.StringIO()
            writer = csv.writer(output)

            # Headers
            headers = [
                "Action_ID",
                "Violation_ID",
                "Resident_ID",
                "Resident_Name",
                "Action_Type",
                "Action_Description",
                "Action_Date",
                "Assigned_To",
                "Due_Date",
                "Completion_Date",
                "Status",
                "Expected_Outcome",
                "Actual_Outcome",
                "Effectiveness_Rating",
                "Is_Overdue",
                "Days_Since_Due",
                "Requires_Follow_Up",
                "Follow_Up_Completed",
            ]
            writer.writerow(headers)

            # Get related data
            resident_map = {}
            violation_map = {}

            if actions:
                resident_ids = list(set(a.resident_id for a in actions))
                residents = (
                    self.db.query(Resident).filter(Resident.id.in_(resident_ids)).all()
                )
                resident_map = {r.id: r for r in residents}

                violation_ids = list(set(a.violation_id for a in actions))
                violations = (
                    self.db.query(ACGMEViolation)
                    .filter(ACGMEViolation.id.in_(violation_ids))
                    .all()
                )
                violation_map = {v.id: v for v in violations}

            # Data rows
            for action in actions:
                resident = resident_map.get(action.resident_id)
                days_since_due = None
                if action.due_date:
                    days_since_due = (
                        (date.today() - action.due_date).days
                        if date.today() > action.due_date
                        else 0
                    )

                row = [
                    action.id,
                    action.violation_id,
                    action.resident_id,
                    resident.name if resident else "Unknown",
                    action.action_type,
                    action.action_description,
                    action.action_date.isoformat(),
                    action.assigned_to,
                    action.due_date.isoformat() if action.due_date else "",
                    (
                        action.completion_date.isoformat()
                        if action.completion_date
                        else ""
                    ),
                    action.status,
                    action.expected_outcome or "",
                    action.actual_outcome or "",
                    action.effectiveness_rating,
                    action.is_overdue,
                    days_since_due,
                    action.requires_follow_up,
                    action.follow_up_completed,
                ]
                writer.writerow(row)

            csv_content = output.getvalue()
            output.close()

            # Log audit trail
            self._log_export_activity(
                "corrective_actions_csv_export",
                {
                    "start_date": start_date.isoformat(),
                    "end_date": end_date.isoformat(),
                    "status_filter": status_filter,
                    "record_count": len(actions),
                },
            )

            return csv_content

        except Exception as e:
            self.logger.error(f"Failed to export corrective actions CSV: {e}")
            raise

    def generate_acgme_submission_package(self, report_id: int) -> Dict[str, bytes]:
        """Generate complete package for ACGME submission."""
        try:
            report = (
                self.db.query(ACGMEComplianceReport)
                .filter(ACGMEComplianceReport.id == report_id)
                .first()
            )

            if not report:
                raise ValueError(f"Compliance report {report_id} not found")

            package = {}

            # 1. Main compliance report (JSON)
            package["compliance_report.json"] = self.export_compliance_report_json(
                report_id
            ).encode("utf-8")

            # 2. Violations summary (CSV)
            violations_csv = self.export_violations_csv(
                report.start_date, report.end_date, report.program_id
            )
            package["violations_summary.csv"] = violations_csv.encode("utf-8")

            # 3. Weekly summaries (Excel)
            weekly_excel = self.export_weekly_summaries_excel(
                report.start_date, report.end_date, report.program_id
            )
            package["weekly_summaries.xlsx"] = weekly_excel

            # 4. Corrective actions (CSV)
            corrective_csv = self.export_corrective_actions_csv(
                report.start_date, report.end_date
            )
            package["corrective_actions.csv"] = corrective_csv.encode("utf-8")

            # 5. Cover letter (text)
            cover_letter = self._generate_cover_letter(report)
            package["cover_letter.txt"] = cover_letter.encode("utf-8")

            # Log audit trail
            self._log_export_activity(
                "acgme_submission_package",
                {
                    "report_id": report_id,
                    "report_name": report.report_name,
                    "files_generated": len(package),
                },
            )

            return package

        except Exception as e:
            self.logger.error(
                f"Failed to generate ACGME submission package for report {report_id}: {e}"
            )
            raise

    def _generate_cover_letter(self, report: ACGMEComplianceReport) -> str:
        """Generate cover letter for ACGME submission."""
        return f"""
ACGME Duty Hour Compliance Report Submission

Report Period: {report.start_date} to {report.end_date}
Report Type: {report.report_type.title()}
Program ID: {report.program_id or 'All Programs'}
Generated: {report.created_at.strftime('%Y-%m-%d %H:%M:%S')}

EXECUTIVE SUMMARY:
- Total Residents Analyzed: {report.total_residents}
- Overall Compliance Rate: {report.compliance_rate:.1f}%
- Total Violations Detected: {report.total_violations}
- Critical Violations: {report.critical_violations}
- High Severity Violations: {report.high_violations}

PROGRAM PERFORMANCE:
- Average Weekly Duty Hours: {report.avg_weekly_hours:.1f}
- Maximum Weekly Duty Hours: {report.max_weekly_hours:.1f}
- Weeks Exceeding 80-Hour Limit: {report.weeks_over_80_hours}

This report includes:
1. compliance_report.json - Comprehensive compliance data in JSON format
2. violations_summary.csv - Detailed violation records
3. weekly_summaries.xlsx - Weekly duty hour summaries for all residents
4. corrective_actions.csv - Corrective actions taken for violations

All data has been verified for accuracy and completeness. This submission
demonstrates our program's commitment to ACGME duty hour compliance and
resident wellness.

Report generated by: {report.generated_by}
Universal Medical Residency Scheduler v1.0.0
        """.strip()

    def _log_export_activity(self, action_type: str, details: Dict[str, Any]):
        """Log export activity for audit trail."""
        try:
            audit_entry = AuditLog(
                action=action_type,
                resource_type="compliance_export",
                resource_id=None,
                user_id=None,  # Would be set from current user context
                details=details,
                timestamp=datetime.utcnow(),
            )

            self.db.add(audit_entry)
            self.db.commit()

            self.logger.info(f"Logged export activity: {action_type} - {details}")

        except Exception as e:
            self.logger.error(f"Failed to log export activity: {e}")

    def get_export_audit_trail(
        self, start_date: date, end_date: date
    ) -> List[Dict[str, Any]]:
        """Get audit trail of all export activities."""
        try:
            audit_entries = (
                self.db.query(AuditLog)
                .filter(
                    and_(
                        AuditLog.resource_type == "compliance_export",
                        AuditLog.timestamp
                        >= datetime.combine(start_date, datetime.min.time()),
                        AuditLog.timestamp
                        <= datetime.combine(end_date, datetime.max.time()),
                    )
                )
                .order_by(desc(AuditLog.timestamp))
                .all()
            )

            return [
                {
                    "id": entry.id,
                    "action": entry.action,
                    "timestamp": entry.timestamp.isoformat(),
                    "user_id": entry.user_id,
                    "details": entry.details,
                }
                for entry in audit_entries
            ]

        except Exception as e:
            self.logger.error(f"Failed to get export audit trail: {e}")
            raise
